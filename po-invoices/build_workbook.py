#!/usr/bin/env python3
"""Build pos_invoices_by_brand.xlsx from pos_invoices_sorted_by_brand.csv.

One sheet per brand (Marah, SONDOS, Wadi Halfa) listing that brand's POs in
a column, plus a Summary sheet with per-month commission per brand.

Everything computed lives in formulas: per-PO commission is
=ROUND(received * rate_cell, 2), totals are SUMs, and the Summary pulls from
the brand sheets with SUMIFS on the order-date column.
"""

import csv
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRANDS = ["Marah", "SONDOS", "Wadi Halfa"]
RATE = {"Marah": 0.06, "SONDOS": 0.06, "Wadi Halfa": 0.04}

ARIAL = "Arial"
F_TITLE = Font(name=ARIAL, size=13, bold=True, color="1F3864")
F_NOTE = Font(name=ARIAL, size=9, italic=True, color="808080")
F_HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=ARIAL, size=10)
F_INPUT = Font(name=ARIAL, size=10, bold=True, color="0000FF")
F_TOTAL = Font(name=ARIAL, size=10, bold=True)
FILL_HDR = PatternFill("solid", fgColor="1F4E79")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
B_BODY = Border(bottom=THIN)
B_TOTAL = Border(top=Side(style="double", color="1F4E79"))
OUTCOME_STYLE = {
    "Delivered": ("006100", "C6EFCE"),
    "Shortage": ("9C5700", "FFEB9C"),
    "Cancelled": ("9C0006", "FFC7CE"),
    "Rejected": ("9C0006", "FFC7CE"),
    "Unfulfilled": ("3F3F3F", "EDEDED"),
}

NF_SAR = "#,##0.00"
NF_INT = "#,##0"
NF_PCT = "0%"
NF_DATE = "m/d/yyyy"
NF_MONTH = "mmm yyyy"

HEADERS = ["PO Number", "Channel", "Order Date", "Delivered Date", "FC", "Units",
           "Requested (SAR)", "Received (SAR)", "Accept %", "Fulfill %", "Outcome",
           "Commission (SAR)"]
WIDTHS = [16, 13, 11.5, 13, 7, 8, 14, 14, 9, 9, 11.5, 15]


def parse_mdy(s: str) -> date:
    m, d, y = (int(x) for x in s.split("/"))
    return date(y, m, d)


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_brand_sheet(ws, brand, rows):
    for i, w in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = f"{brand} — Vendor Purchase Orders & Commission"
    ws["A1"].font = F_TITLE
    ws["A2"] = "Commission rate (of received):"
    ws["A2"].font = F_BODY
    ws["B2"] = RATE[brand]
    ws["B2"].font = F_INPUT
    ws["B2"].fill = FILL_INPUT
    ws["B2"].number_format = NF_PCT
    ws["D2"] = "Commission per PO = ROUND(Received × rate, 2). Source: vendor PO dashboard export; delivered dates from Nasam."
    ws["D2"].font = F_NOTE

    hdr_row = 4
    for c, h in enumerate(HEADERS, start=1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, len(HEADERS))

    first = hdr_row + 1
    for i, r in enumerate(rows):
        rr = first + i
        values = [
            r["po_number"], r["channel"], parse_mdy(r["order_date"]),
            parse_mdy(r["delivered_date"]), r["fulfillment_center"], int(r["units"]),
            float(r["requested_sar"]), float(r["received_sar"]),
            int(r["accept_pct"]) / 100, int(r["fulfill_pct"]) / 100, r["outcome"],
            f"=ROUND(H{rr}*$B$2,2)",
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.font = F_BODY
            cell.border = B_BODY
        ws.cell(row=rr, column=3).number_format = NF_DATE
        ws.cell(row=rr, column=4).number_format = NF_DATE
        ws.cell(row=rr, column=6).number_format = NF_INT
        for c in (7, 8, 12):
            ws.cell(row=rr, column=c).number_format = NF_SAR
        for c in (9, 10):
            ws.cell(row=rr, column=c).number_format = NF_PCT
        oc = ws.cell(row=rr, column=11)
        font_color, fill_color = OUTCOME_STYLE[r["outcome"]]
        oc.font = Font(name=ARIAL, size=10, color=font_color)
        oc.fill = PatternFill("solid", fgColor=fill_color)
        oc.alignment = Alignment(horizontal="center")

    last = first + len(rows) - 1
    trow = last + 1
    ws.cell(row=trow, column=1, value="Total")
    ws.cell(row=trow, column=6, value=f"=SUM(F{first}:F{last})").number_format = NF_INT
    for col in ("G", "H", "L"):
        c = ws[f"{col}{trow}"]
        c.value = f"=SUM({col}{first}:{col}{last})"
        c.number_format = NF_SAR
    for c in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=trow, column=c)
        cell.font = F_TOTAL
        cell.border = B_TOTAL

    ws.auto_filter.ref = f"A{hdr_row}:L{last}"
    ws.freeze_panes = f"A{first}"
    return first, last


def month_starts():
    ms, y, m = [], 2025, 9
    for _ in range(12):
        ms.append(date(y, m, 1))
        m += 1
        if m == 13:
            y, m = y + 1, 1
    return ms


def build_summary(ws, ranges):
    for i, w in enumerate([12, 13, 13, 13, 16, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = "PO Commission Summary — Sep 2025 to Aug 2026"
    ws["A1"].font = F_TITLE
    ws["A2"] = "Grouped by PO order month. All figures pulled live from the brand sheets (SAR)."
    ws["A2"].font = F_NOTE

    hdr_row = 4
    for c, h in enumerate(["Month", "Marah", "SONDOS", "Wadi Halfa",
                           "Total Commission", "Total Received"], start=1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, 6)

    first = hdr_row + 1
    months = month_starts()
    for i, ms in enumerate(months):
        rr = first + i
        mc = ws.cell(row=rr, column=1, value=ms)
        mc.number_format = NF_MONTH
        mc.font = F_BODY
        mc.border = B_BODY
        for c, brand in enumerate(BRANDS, start=2):
            sheet = f"'{brand}'" if " " in brand else brand
            f, l = ranges[brand]
            cell = ws.cell(row=rr, column=c)
            cell.value = (f"=SUMIFS({sheet}!$L${f}:$L${l},{sheet}!$C${f}:$C${l},"
                          f'">="&$A{rr},{sheet}!$C${f}:$C${l},"<"&EDATE($A{rr},1))')
            cell.number_format = NF_SAR
            cell.font = F_BODY
            cell.border = B_BODY
        tc = ws.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})")
        rec_parts = []
        for brand in BRANDS:
            sheet = f"'{brand}'" if " " in brand else brand
            f, l = ranges[brand]
            rec_parts.append(f"SUMIFS({sheet}!$H${f}:$H${l},{sheet}!$C${f}:$C${l},"
                             f'">="&$A{rr},{sheet}!$C${f}:$C${l},"<"&EDATE($A{rr},1))')
        rc = ws.cell(row=rr, column=6, value="=" + "+".join(rec_parts))
        for cell in (tc, rc):
            cell.number_format = NF_SAR
            cell.font = F_BODY
            cell.border = B_BODY

    last = first + len(months) - 1
    trow = last + 1
    ws.cell(row=trow, column=1, value="Total")
    for c in range(2, 7):
        col = get_column_letter(c)
        ws.cell(row=trow, column=c, value=f"=SUM({col}{first}:{col}{last})").number_format = NF_SAR
    for c in range(1, 7):
        cell = ws.cell(row=trow, column=c)
        cell.font = F_TOTAL
        cell.border = B_TOTAL

    ws.freeze_panes = f"A{first}"


def main() -> None:
    with open("pos_invoices_sorted_by_brand.csv", newline="") as fh:
        all_rows = list(csv.DictReader(fh))
    assert len(all_rows) == 199

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    ranges = {}
    for brand in BRANDS:
        rows = [r for r in all_rows if r["brand"] == brand]
        ws = wb.create_sheet(brand)
        ranges[brand] = build_brand_sheet(ws, brand, rows)

    build_summary(summary, ranges)
    wb.save("pos_invoices_by_brand.xlsx")
    print("OK: wrote pos_invoices_by_brand.xlsx",
          {b: ranges[b][1] - ranges[b][0] + 1 for b in BRANDS})


if __name__ == "__main__":
    main()
