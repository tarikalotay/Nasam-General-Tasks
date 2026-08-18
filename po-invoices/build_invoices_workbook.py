#!/usr/bin/env python3
"""Build commission_invoices.xlsx — the accountant-facing workbook.

One invoice per (brand, month); each PO with received > 0 is one line inside
its invoice, carrying only: PO number, month, received SAR, commission %,
commission SAR. Customer is Alaris Group for Marah/SONDOS, Wadi Halfa for
Wadi Halfa. The Summary sheet lists every invoice with its PO count, pulling
amounts live from the brand sheets.
"""

import csv
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRANDS = ["Marah", "SONDOS", "Wadi Halfa"]
CUSTOMER = {"Marah": "Alaris Group", "SONDOS": "Alaris Group", "Wadi Halfa": "Wadi Halfa"}
RATE = {"Marah": 0.06, "SONDOS": 0.06, "Wadi Halfa": 0.04}
REF_TAG = {"Marah": "MARAH", "SONDOS": "SONDOS", "Wadi Halfa": "WADIHALFA"}

ARIAL = "Arial"
F_TITLE = Font(name=ARIAL, size=13, bold=True, color="1F3864")
F_NOTE = Font(name=ARIAL, size=9, italic=True, color="808080")
F_HDR = Font(name=ARIAL, size=10, bold=True, color="FFFFFF")
F_BODY = Font(name=ARIAL, size=10)
F_INPUT = Font(name=ARIAL, size=10, bold=True, color="0000FF")
F_TOTAL = Font(name=ARIAL, size=10, bold=True)
F_SECTION = Font(name=ARIAL, size=10, bold=True, color="1F4E79")
FILL_HDR = PatternFill("solid", fgColor="1F4E79")
FILL_SECTION = PatternFill("solid", fgColor="DDEBF7")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="BFBFBF")
B_BODY = Border(bottom=THIN)
B_TOTAL = Border(top=Side(style="thin", color="1F4E79"))

NF_SAR = "#,##0.00"
NF_PCT = "0%"
NF_MONTH = "mmm yyyy"


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HDR
        cell.fill = FILL_HDR
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_brand_sheet(ws, brand, invoices):
    """invoices: list of (month_date, ref, [rows]) in chronological order.

    Returns per-invoice metadata for the Summary sheet:
    (ref, month_date, first_po_row, last_po_row, total_row).
    """
    for i, w in enumerate([18, 11, 14, 13, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = f"{brand} — Commission Invoices — Customer: {CUSTOMER[brand]}"
    ws["A1"].font = F_TITLE
    ws["A2"] = "Commission rate (of received):"
    ws["A2"].font = F_BODY
    ws["B2"] = RATE[brand]
    ws["B2"].font = F_INPUT
    ws["B2"].fill = FILL_INPUT
    ws["B2"].number_format = NF_PCT
    ws["C2"] = "Commission per PO = ROUND(Received × rate, 2)."
    ws["C2"].font = F_NOTE

    hdr_row = 4
    for c, h in enumerate(["PO Number", "Month", "Received (SAR)",
                           "Commission %", "Commission (SAR)"], start=1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, 5)
    ws.freeze_panes = f"A{hdr_row + 1}"

    meta = []
    rr = hdr_row + 1
    for month, ref, rows in invoices:
        sec = ws.cell(row=rr, column=1,
                      value=f"{ref}  ·  {CUSTOMER[brand]}  ·  {month.strftime('%b %Y')}")
        for c in range(1, 6):
            cell = ws.cell(row=rr, column=c)
            cell.fill = FILL_SECTION
            cell.font = F_SECTION
        rr += 1

        first = rr
        for r in rows:
            ws.cell(row=rr, column=1, value=r["po_number"])
            mc = ws.cell(row=rr, column=2, value=month)
            mc.number_format = NF_MONTH
            ws.cell(row=rr, column=3, value=float(r["received_sar"])).number_format = NF_SAR
            ws.cell(row=rr, column=4, value="=$B$2").number_format = NF_PCT
            ws.cell(row=rr, column=5, value=f"=ROUND(C{rr}*$B$2,2)").number_format = NF_SAR
            for c in range(1, 6):
                cell = ws.cell(row=rr, column=c)
                cell.font = F_BODY
                cell.border = B_BODY
            rr += 1
        last = rr - 1

        ws.cell(row=rr, column=1, value="Invoice total")
        ws.cell(row=rr, column=3, value=f"=SUM(C{first}:C{last})").number_format = NF_SAR
        ws.cell(row=rr, column=5, value=f"=SUM(E{first}:E{last})").number_format = NF_SAR
        for c in range(1, 6):
            cell = ws.cell(row=rr, column=c)
            cell.font = F_TOTAL
            cell.border = B_TOTAL
        meta.append((ref, month, first, last, rr))
        rr += 2  # blank row between invoices

    return meta


def build_summary(ws, meta_by_brand):
    for i, w in enumerate([24, 14, 12, 10, 7, 14, 15], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws["A1"] = "Commission Invoices — Summary"
    ws["A1"].font = F_TITLE

    n_invoices = sum(len(m) for m in meta_by_brand.values())
    ws["A2"] = "Number of invoices:"
    ws["A2"].font = F_BODY
    ws["B2"].font = F_TOTAL

    hdr_row = 4
    for c, h in enumerate(["Invoice", "Customer", "Brand", "Month", "POs",
                           "Received (SAR)", "Commission (SAR)"], start=1):
        ws.cell(row=hdr_row, column=c, value=h)
    style_header_row(ws, hdr_row, 7)
    ws.freeze_panes = f"A{hdr_row + 1}"

    entries = []
    for brand in BRANDS:
        for ref, month, first, last, trow in meta_by_brand[brand]:
            entries.append((month, brand, ref, first, last, trow))
    entries.sort(key=lambda e: (e[0], e[1]))

    rr = hdr_row + 1
    for month, brand, ref, first, last, trow in entries:
        sheet = f"'{brand}'" if " " in brand else brand
        ws.cell(row=rr, column=1, value=ref)
        ws.cell(row=rr, column=2, value=CUSTOMER[brand])
        ws.cell(row=rr, column=3, value=brand)
        mc = ws.cell(row=rr, column=4, value=month)
        mc.number_format = NF_MONTH
        ws.cell(row=rr, column=5, value=f"=COUNTA({sheet}!$A${first}:$A${last})")
        ws.cell(row=rr, column=6, value=f"={sheet}!$C${trow}").number_format = NF_SAR
        ws.cell(row=rr, column=7, value=f"={sheet}!$E${trow}").number_format = NF_SAR
        for c in range(1, 8):
            cell = ws.cell(row=rr, column=c)
            cell.font = F_BODY
            cell.border = B_BODY
        rr += 1

    first_data = hdr_row + 1
    last_data = rr - 1
    ws["B2"] = f"=COUNTA(A{first_data}:A{last_data})"
    ws.cell(row=rr, column=1, value="Total")
    for c in (5, 6, 7):
        col = get_column_letter(c)
        cell = ws.cell(row=rr, column=c, value=f"=SUM({col}{first_data}:{col}{last_data})")
        cell.number_format = NF_SAR if c > 5 else "0"
    for c in range(1, 8):
        cell = ws.cell(row=rr, column=c)
        cell.font = F_TOTAL
        cell.border = B_TOTAL

    return n_invoices


def main() -> None:
    with open("pos_invoices_sorted_by_brand.csv", newline="") as fh:
        all_rows = [r for r in csv.DictReader(fh) if Decimal(r["received_sar"]) > 0]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"

    meta_by_brand = {}
    n_pos = 0
    for brand in BRANDS:
        rows = [r for r in all_rows if r["brand"] == brand]
        n_pos += len(rows)
        groups = {}
        for r in rows:
            m, _, y = r["order_date"].split("/")
            groups.setdefault(date(int(y), int(m), 1), []).append(r)
        invoices = [(month, f"INV-{REF_TAG[brand]}-{month.year}-{month.month:02d}",
                     groups[month]) for month in sorted(groups)]
        ws = wb.create_sheet(brand)
        meta_by_brand[brand] = build_brand_sheet(ws, brand, invoices)

    n_inv = build_summary(summary, meta_by_brand)
    wb.save("commission_invoices.xlsx")
    print(f"OK: wrote commission_invoices.xlsx — {n_inv} invoices, {n_pos} PO lines",
          {b: len(meta_by_brand[b]) for b in BRANDS})


if __name__ == "__main__":
    main()
