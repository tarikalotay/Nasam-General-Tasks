# Nasam Revenue View — weekly builder (canonical). Run from this directory: python3 build.py
import os
HERE=os.path.dirname(os.path.abspath(__file__))
import re, csv
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import rev5, old_gmv, add6, wafeq_api, po_plat

F="Arial"
BLACK=Font(name=F,size=10); BOLD=Font(name=F,size=10,bold=True)
HDR=Font(name=F,size=10,bold=True,color="FFFFFF")
HFILL=PatternFill("solid",fgColor="16323C"); GREY=PatternFill("solid",fgColor="F2EFE8")
SALES_F=PatternFill("solid",fgColor="E9F1F9"); BILLED_F=PatternFill("solid",fgColor="E7F2EA"); TOINV_F=PatternFill("solid",fgColor="FBF3DC")
SALLA_SHARE=0.15
SAR='#,##0.00;(#,##0.00);-'; SAR0='#,##0;(#,##0);-'
CUT='2025-11'
M=[f"20{y}-{m:02d}" for y,m in [(25,11),(25,12)]+[(26,i) for i in range(1,9)]]
MNAMES={1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
def mlbl(m): return f"{MNAMES[int(m[5:7])]} {m[:4]}"
ML=[mlbl(m) for m in M]
SONBOL="Sonbol"
SONBOL_OLD="سنبل - المتجر الإلكتروني - Sonbol"

SRC_WAFEQ="Wafeq invoice export"
SRC_PLAT="Nasam platform — revenue view"
SRC_ORD="Nasam platform — orders view"
SRC_PO="Nasam platform — closed POs"
SRC_CARD="Rate card (Tarik)"
SRC_SALLA="Salla Partners export"

CL=[
 ("Alfaris Group",["شركة مجموعة الفارس الصناعية"],"Active","",["SONDOS","Marah","Sense","Arabesque"]),
 ("Inivita",["مؤسسة صحة ونكهة للتجارة"],"Active","",["Invita"]),
 ("Nokush",["مؤسسة نقوش القهوة للتجارة"],"Active","",["Nokush"]),
 ("Wadi Halfa",["شركة وادي حلفا للتجارة"],"Active","",["Wadi Halfa"]),
 ("Dar Sonbol",["شركة دار سنبل للتجاره"],"Active","",[SONBOL]),
 ("Two United",["شركة تو المتحدة للتجارة - Two"],"Churned","2026-05",["Bonita","brio","circles","Feel","labelle","Reefi"]),
 ("Rimath",["شركة البخور الذكي - IOUD","شركة البخور الذكي - Thannah","شركة البخور الذكي - Oso spray"],"Churned","2026-06",["iOUD","Thannah","O'SO"]),
 ("Ancy & Shaya",["شركة التكامل الثلاثي للتجارة"],"Churned","2026-05",["Ancy","Shaya"]),
 ("Gamesir",["شركة قمة التطورات المستقبلية التجارية"],"Churned","2026-02",[]),
 ("Safwat Aljouf",["صفوة الجوف الزراعية التجارية"],"Churned","2026-02",["Safwat Aljawf"]),
 ("Reetal",["مؤسسة محمد صالح محمد العشري للتجارة"],"Churned","2026-06",["Reetal"]),
 ("House Mart",["شركة الريادية الابتكارية للتجارة"],"Churned","2026-04",[]),
]
EXCL_CONTACT={"Cloud Shelf":"x","مؤسسة قيادة آمنة للتجارة":"x"}
C2K={}
for k,cons,*_ in CL:
    for c in cons: C2K[c]=k
CONTACT_BRAND={"شركة البخور الذكي - IOUD":"iOUD","شركة البخور الذكي - Thannah":"Thannah","شركة البخور الذكي - Oso spray":"O'SO"}
B2C={}; K2B={k:brs for k,_,_,_,brs in CL}; CHURN={k:chn for k,_,_,chn,_ in CL}
for k,_,_,_,brs in CL:
    for b in brs: B2C[b]=k
B2C[SONBOL_OLD]="Dar Sonbol"
DISP={SONBOL:"Sonbol (سنبل)",SONBOL_OLD:"Sonbol (سنبل)"}
def disp(b): return DISP.get(b,b)
def ckey(contact):
    if contact in C2K: return C2K[contact]
    s=str(contact or '')
    for c,k in C2K.items():
        if s[:15] and (c.startswith(s[:15]) or s[:12] in c): return k
    return None
SAAS={("SONDOS","Salla"),("Marah","Salla"),("Sense","Salla"),("Arabesque","Salla"),
      (SONBOL,"Salla"),("Ancy","Noon"),("Shaya","Noon")}
def model(b,ch):
    return "SaaS" if (b,ch) in SAAS else "FAM"

# SaaS app subscriptions — Salla Partners export 18 Aug 2026 (منصة نسم)
SAAS_SUBS=[ # store, client, merchant#, plan, start, end, list, net ex-VAT, paid inc-VAT, coupon, status
 ("محمصة نقوش (Nokush)","Nokush","317767866","Annual","2026-05-14","2027-05-14",3999.00,999.75,1149.71,"SPNS75 (−75%)","Paid — active"),
 ("SONDOS","Alfaris Group","1052755076","Annual","2026-08-09","2027-08-09",3999.00,999.75,1149.71,"SPNS75 (−75%)","Paid — active"),
 ("Sonbol (سنبل)","Dar Sonbol","773002475","Annual","2026-08-13","2026-08-20",None,0.0,0.0,"—","Trial until 20 Aug"),
 ("Pearly touch","— new lead","458981891","Monthly","2026-08-15","2026-08-22",None,0.0,0.0,"—","Trial until 22 Aug"),
 ("شركة جرافيتي ريزن","— new lead","825450288","Monthly","2026-08-12","2026-08-19",None,0.0,0.0,"—","Trial until 19 Aug"),
]
saas_rev=defaultdict(float)
for s in SAAS_SUBS:
    if s[7]>0: saas_rev[s[4][:7]]+=s[7]
saas_total=sum(saas_rev.values())

# ---------- parse Wafeq ----------
AR_M={'يناير':1,'فبراير':2,'مارس':3,'أبريل':4,'ابريل':4,'مايو':5,'يونيو':6,'يوليو':7,'أغسطس':8,'اغسطس':8,'سبتمبر':9,'أكتوبر':10,'اكتوبر':10,'نوفمبر':11,'ديسمبر':12}
EN_M={'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}
def parse_ref(ref,fallback):
    s=str(ref or '')
    m=re.search(r'(20\d\d)',s)
    if m:
        y=m.group(1)
        for name,mm in EN_M.items():
            if name in s.lower(): return f"{y}-{mm:02d}"
        for name,mm in AR_M.items():
            if name in s: return f"{y}-{mm:02d}"
    return fallback
def fnum(x):
    try: return float(x)
    except: return 0.0
CH_AR=[('أمازون للتجزئة','Amazon Retail'),('أمازون','Amazon'),('ترنديول','Trendyol'),('سلة','Salla'),('نون','Noon'),('منصات التجارة السريعة','Quick-commerce')]
def stream(item,desc,client):
    it=str(item or ''); de=str(desc or '')
    if 'كلاود شيلف' in de or 'كلاود شيلف' in it: return ('PassThrough',None)
    if 'PO#' in de or 'PO #' in de: return ('Commission','Ninja Retail' if 'Ninja' in de or 'نينجا' in de else 'Amazon Retail')
    if 'رسوم شهرية' in it: return ('Monthly fee',None)
    if it.startswith('عمولة مبيعات'):
        for a,e in CH_AR:
            if a in it: return ('Commission',e)
        return ('Commission','Other')
    if it.startswith('فتح حساب'): return ('Setup',None)
    if any(x in de for x in ['إدراج المنتجات','تحسين وصف','تصميم','تعديل بيانات']): return ('Setup',None)
    if 'عمولة' in de: return ('Commission','Other')
    if any(x in de for x in ['تسويق','تصوير']): return ('Other',None)
    if 'فتح' in de or 'تسجيل' in de: return ('Setup',None)
    if client=='House Mart': return ('Setup',None)
    return ('Other',None)

def wafeq_lines():
    if wafeq_api.available():
        for L in wafeq_api.load(): yield L
        return
    ws0=load_workbook(os.path.join(HERE,'sources','wafeq_invoices_2026-08-17.xlsx'), data_only=True)['Sheet1']
    seen=set()
    for r in ws0.iter_rows(min_row=2, values_only=True):
        num=r[1]; first=num not in seen; seen.add(num)
        yield dict(num=num,date=str(r[2])[:10],contact=r[5],amount=fnum(r[8]),balance=fnum(r[10]),
                   status=str(r[0]),disc=fnum(r[26]),ref=r[32],item=str(r[12] or ''),desc=str(r[15] or ''),lam=fnum(r[23]),first=first)
SRC_MODE='Wafeq API' if wafeq_api.available() else 'manual xlsx export'
print('billing source:',SRC_MODE)
invmeta={}
fee=defaultdict(float); setup=defaultdict(float); other=defaultdict(float); disc=defaultdict(float); pth=defaultdict(float)
comm=defaultdict(float); comm_brand=defaultdict(float); rate_obs=defaultdict(set)
for L in wafeq_lines():
    num=L['num']; contact=L['contact']; item=L['item']; desc=L['desc']; lam=L['lam']
    month=parse_ref(L['ref'], L['date'][:7])
    if contact in EXCL_CONTACT:
        era='Excluded'; ck='(excluded)'
    else:
        ck=ckey(contact) or f"UNRESOLVED:{contact}"
        era='Current' if month>=CUT else 'Old'
    if num not in invmeta:
        invmeta[num]={'amount':L['amount'],'client':ck,'era':era,'month':month,'disc':L['disc']}
    if era!='Current': continue
    st,ch=stream(item,desc,ck)
    if st=='PassThrough': pth[(ck,month)]+=lam
    elif st=='Monthly fee': fee[(ck,month)]+=lam
    elif st=='Setup': setup[(ck,month)]+=lam
    elif st=='Other': other[(ck,month)]+=lam
    elif st=='Commission':
        comm[(ck,ch,month)]+=lam
        b=CONTACT_BRAND.get(contact)
        if b: comm_brand[(ck,b,ch,month)]+=lam
        mm=re.search(r'\((\d+(?:[.,]\d+)?)٪\)', str(desc or ''))
        if mm: rate_obs[(ck,ch)].add(mm.group(1).replace(',','.'))
for num,v in invmeta.items():
    if v['disc']>0.01 and v['era']=='Current':
        disc[(v['client'],v['month'])]-=v['disc']
unres=sorted({v['client'] for v in invmeta.values() if str(v['client']).startswith('UNRESOLVED')})
assert not unres, f"unresolved contacts: {unres}"
_x=os.path.join(HERE,'sources','wafeq_invoices_2026-08-17.xlsx')
if wafeq_api.available() and os.path.exists(_x):
    _w=load_workbook(_x, data_only=True)['Sheet1']; _seen=set(); _xt={}
    for _r in _w.iter_rows(min_row=2, values_only=True):
        if _r[1] in _seen: continue
        _seen.add(_r[1]); _xt[_r[1]]=fnum(_r[8])
    _d=[n for n in _xt if abs(_xt[n]-(invmeta.get(n,{}).get('amount',0)))>0.02]
    print(f"API-vs-export reconciliation: {len(_xt)} shared invoices, {len(_d)} mismatches"+(f" -> {_d[:5]}" if _d else ""))
allk=[k for k,*_ in CL]
billed_wafeq=defaultdict(float)
for num,v in invmeta.items():
    if v['era']=='Current': billed_wafeq[(v['client'],v['month'])]+=v['amount']
strip_total=sum(pth.values())
billed=defaultdict(float)
for kk,vv in billed_wafeq.items(): billed[kk]=vv
for (ck,m),v in pth.items(): billed[(ck,m)]-=v

# ---------- retail: closed POs from the platform (values live since 20 Aug), netted vs Wafeq ----------
acc=defaultdict(lambda: defaultdict(float)); acc_val=defaultdict(lambda: defaultdict(float)); acc_rate={}
plat=defaultdict(float)
for b,ch,ym,rec in po_plat.PO_CLOSED:
    k=B2C[b]; rate=po_plat.RETAIL_RATE[b]
    acc_val[(k,b,ch)][ym]+=rec
    plat[(k,b,ch,ym)]+=rec*rate
    acc_rate[(k,b,ch)]=rate*100
plat_total=sum(plat.values())
billed_retail=defaultdict(float)
for (k2,ch2,m2),v in comm.items():
    if ch2 in ('Amazon Retail','Ninja Retail'): billed_retail[(k2,ch2)]+=v
over=0.0
for (k2,ch2),bill in billed_retail.items():
    keys=sorted([kk for kk in plat if kk[0]==k2 and kk[2]==ch2], key=lambda x:x[3])
    rem=bill
    for kk in keys:
        take=min(plat[kk],rem); plat[kk]-=take; rem-=take
        if rem<=0.005: break
    if rem>0.005: over+=rem
for (k2,b,ch2,m2),v in plat.items():
    if v>0.005: acc[(k2,b,ch2)][m2]+=v
acc_total=sum(v for d in acc.values() for v in d.values())
acc_val_total=sum(v for d in acc_val.values() for v in d.values())
billed_retail_total=sum(billed_retail.values())
print(f"retail: platform commission {plat_total:,.2f} | billed in Wafeq {billed_retail_total:,.2f} | net to invoice {acc_total:,.2f}"+(f" | over-billed vs platform {over:,.2f}" if over>0.005 else ""))

# ---------- GMV ----------
gmv=defaultdict(float); gmv_src={}
def covered(b,ym):
    k=B2C.get(b)
    if not k or ym<CUT: return False
    chn=CHURN.get(k)
    if chn and ym>chn: return False
    return True
for b,c,ym,g,o in rev5.REV_ACT:
    if b==SONBOL: continue
    if covered(b,ym): gmv[(b,c,ym)]+=g; gmv_src[(b,c)]=SRC_PLAT
for b,c,ym,o,g in old_gmv.OLD_M:
    if covered(b,ym): gmv[(b,c,ym)]+=g; gmv_src[(b,c)]=SRC_PLAT
for b,c,ym,g,o in add6.TWO_M:
    if covered(b,ym): gmv[(b,c,ym)]+=g; gmv_src[(b,c)]=SRC_ORD
gmv[(SONBOL,"Salla","2026-08")]=add6.SONBOL_FROM_INTEGRATION[0]
gmv_src[(SONBOL,"Salla")]=SRC_PLAT+" (from 13 Aug 2026)"

CH_PLAT={'Amazon':['Amazon'],'Trendyol':['Trendyol'],'Salla':['Salla'],'Noon':['Noon'],
         'Quick-commerce':['Jahez','Hungerstation','NoonFood'],'Amazon Retail':[],'Other':[]}
P2W={'Amazon':'Amazon','Trendyol':'Trendyol','Salla':'Salla','Noon':'Noon','Jahez':'Quick-commerce','Hungerstation':'Quick-commerce','NoonFood':'Quick-commerce'}

alloc=defaultdict(lambda: defaultdict(float)); alloc_note=set()
for (ck,ch,month),amt in comm.items():
    exact=[(b,v) for (c2,b,c3,m2),v in comm_brand.items() if c2==ck and c3==ch and m2==month]
    if exact:
        for b,v in exact: alloc[(ck,b,ch)][month]+=v
        continue
    brs=K2B.get(ck,[]); plat=CH_PLAT.get(ch,[])
    if brs and (len(brs)==1 or not plat):
        alloc[(ck,brs[0] if len(brs)==1 else '— all brands',ch)][month]+=amt
        continue
    if brs and plat:
        shares={b:sum(gmv.get((b,pc,month),0.0) for pc in plat) for b in brs}
        tot=sum(shares.values())
        if tot>0:
            live=[b for b,s in shares.items() if s>0]
            for b in live: alloc[(ck,b,ch)][month]+=amt*shares[b]/tot
            if len(live)>1: alloc_note.add((ck,ch))
            continue
    alloc[(ck,'— (no platform brand)',ch)][month]+=amt
for (ck,ch,month),amt in comm.items():
    got=sum(mm.get(month,0.0) for (c2,b,c3),mm in alloc.items() if c2==ck and c3==ch)
    assert abs(got-amt)<0.02, f"alloc mismatch {ck}/{ch}/{month}"
print("brand allocation preserves billed commission: OK")

def rate_str(ck,ch):
    v=sorted(rate_obs.get((ck,ch),set()))
    return " / ".join(f"{x}%" for x in v) if v else ""

wb=Workbook()
def hdr(ws,hs,r=1,widths=None):
    for i,h in enumerate(hs,1):
        c=ws.cell(row=r,column=i,value=h); c.font=HDR; c.fill=HFILL
        c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    if widths:
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes=ws.cell(row=r+1,column=1)
    ws.sheet_properties.outlinePr.summaryBelow=False
def put(ws,r,c,v,font=BLACK,fmt=None,fill=None):
    cell=ws.cell(row=r,column=c,value=v); cell.font=font
    if fmt: cell.number_format=fmt
    if fill: cell.fill=fill
    return cell
def olvl(ws,r,lvl):
    if lvl>0: ws.row_dimensions[r].outline_level=lvl

# ---------- 0 ReadMe ----------
w=wb.active; w.title="0 ReadMe"
w.column_dimensions['B'].width=26; w.column_dimensions['C'].width=112
rows=[("Nasam Revenue View — weekly run · 23 Aug 2026",""),
("This run","First automated Sunday run. Billing source switched to the WAFEQ API (reconciled against the 17-Aug manual export before switching). Retail reads closed POs directly from the platform (values live since 20 Aug), netted against retail invoices already raised in Wafeq. Sonbol was renamed in the platform ('Sonbol') and its from-integration sales refreshed."),
("Sources","Wafeq API snapshot 23 Aug (160 invoices) · platform revenue + purchase orders pulled 23 Aug · churned-brand snapshots 15 Aug (the read layer does not serve deactivated brands) · Salla Partners export 18 Aug (manual) · rate card 17 Aug."),
("Reporting rules","Window Nov 2025+ (current model) · GMV post-Nasam only (Sonbol from 13 Aug 2026) · post-churn months excluded · SaaS brand-channels 0% · SaaS subscriptions net of Salla 15% · retail commission on the RECEIVED value of closed POs · pass-throughs (3,355.01) excluded · all figures SAR ex-VAT."),
("Weekly update","Automatic: Wafeq API, platform sales, platform POs, rebuild + verification. Manual: only the Salla Partners subscriptions export — share it whenever it changes."),]
for i,(a,b) in enumerate(rows, start=2):
    w.cell(row=i,column=2,value=a).font=BOLD if i==2 else BLACK
    c=w.cell(row=i,column=3,value=b); c.font=BLACK; c.alignment=Alignment(wrap_text=True,vertical="top")
    w.row_dimensions[i].height=None if i==2 else 52

def month_row(ws,r,label,getter,src,font=BLACK,fill=None,fmt=SAR0,lvl=0):
    put(ws,r,2,label,font,fill=fill); tot=0.0
    for j,m in enumerate(M):
        v=getter(m); tot+=v
        put(ws,r,3+j,round(v,2) if abs(v)>0.005 else None,font,fmt,fill)
    put(ws,r,3+len(M),round(tot,2),font,fmt,fill)
    put(ws,r,4+len(M),src,BLACK)
    olvl(ws,r,lvl)
    return tot

# sales tree data: client -> brand -> channel -> {month: value}; retail received merged in
tree=defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(float))))
tree_src={}
for (b,c,m),v in gmv.items():
    k=B2C[b]; tree[k][b][c][m]+=v; tree_src[(k,b,c)]=gmv_src.get((b,c),SRC_PLAT)
for (k,b,ch),mm in acc_val.items():
    for m,v in mm.items(): tree[k][b][ch][m]+=v
    tree_src[(k,b,ch)]=SRC_PO+" — received value"

# ---------- 1 Revenue Summary ----------
w1=wb.create_sheet("1 Revenue Summary")
hdr(w1,["","Item"]+ML+["Total","Source"],widths=[3,34]+[10]*len(M)+[12,46])
r=2
put(w1,r,2,"SALES",BOLD); r+=1
def gmv_by(pred):
    return lambda m: sum(v for (b,c,m2),v in gmv.items() if m2==m and pred(b,c))
man_tot=month_row(w1,r,"Managed (FAM) marketplace GMV",gmv_by(lambda b,c: model(b,c)=="FAM"),SRC_PLAT); r+=1
ret_tot=month_row(w1,r,"Managed retail — received value",lambda m: sum(d.get(m,0.0) for d in acc_val.values()),SRC_PO+" — received value"); r+=1
saas_tot=month_row(w1,r,"SaaS (unmanaged) GMV",gmv_by(lambda b,c: model(b,c)=="SaaS"),SRC_PLAT); r+=1
all_tot=month_row(w1,r,"Total sales",lambda m: sum(v for (b,c,m2),v in gmv.items() if m2==m)+sum(d.get(m,0.0) for d in acc_val.values()),"Sum of the three rows above",BOLD,GREY); r+=1
sales_hdr=r
put(w1,r,2,"SALES BY CLIENT / BRAND / CHANNEL",BOLD); r+=1
for k,cons,st_,chn,brs in CL:
    kb=tree.get(k)
    if not kb: continue
    brands=sorted(kb.keys())
    single_brand=len(brands)==1
    single_all=single_brand and len(kb[brands[0]])==1
    lbl=k+(f"  (churned {chn})" if chn else "")
    if single_all:
        b=brands[0]; c=list(kb[b].keys())[0]
        month_row(w1,r,f"{lbl} — {disp(b)} · {c}",lambda m,bb=b,cc=c,kk=k: tree[kk][bb][cc].get(m,0.0),tree_src[(k,b,c)]); r+=1
        continue
    month_row(w1,r,lbl,lambda m,kk=k: sum(v for bb in tree[kk] for cc in tree[kk][bb] for m2,v in tree[kk][bb][cc].items() if m2==m),"Sum of rows below",BOLD); r+=1
    for b in brands:
        chans=sorted(kb[b].keys())
        if not single_brand:
            if len(chans)==1:
                c=chans[0]
                month_row(w1,r,f"{disp(b)} · {c}",lambda m,bb=b,cc=c,kk=k: tree[kk][bb][cc].get(m,0.0),tree_src[(k,b,c)],lvl=1); r+=1
                continue
            month_row(w1,r,disp(b),lambda m,bb=b,kk=k: sum(v for cc in tree[kk][bb] for m2,v in tree[kk][bb][cc].items() if m2==m),"Sum of rows below",BLACK,lvl=1); r+=1
        for c in chans:
            if not single_brand and len(chans)==1: continue
            month_row(w1,r,f"{disp(b)+' · ' if single_brand else '    '}{c}" if single_brand else f"    {c}",lambda m,bb=b,cc=c,kk=k: tree[kk][bb][cc].get(m,0.0),tree_src[(k,b,c)],lvl=1 if single_brand else 2); r+=1
r+=1
put(w1,r,2,"NASAM REVENUE BY STREAM",BOLD); r+=1
sums={}
sums['Monthly fees']=month_row(w1,r,"Monthly fees",lambda m: sum(fee[(k,m)] for k in allk),SRC_WAFEQ); r+=1
sums['Commission']=month_row(w1,r,"Commission (billed)",lambda m: sum(v for (k,ch,m2),v in comm.items() if m2==m),SRC_WAFEQ); r+=1
sums['Setup']=month_row(w1,r,"Setup / one-time",lambda m: sum(setup[(k,m)] for k in allk),SRC_WAFEQ); r+=1
sums['Other']=month_row(w1,r,"Other",lambda m: sum(other[(k,m)] for k in allk),SRC_WAFEQ); r+=1
sums['Discounts']=month_row(w1,r,"Discounts",lambda m: sum(disc[(k,m)] for k in allk),SRC_WAFEQ); r+=1
tot_billed=month_row(w1,r,"TOTAL BILLED",lambda m: sum(billed[(k,m)] for k in allk),SRC_WAFEQ+" — ties to ledger",BOLD,GREY); r+=1
acc_row=month_row(w1,r,"Retail commission — closed POs, to invoice",lambda m: sum(d.get(m,0.0) for d in acc.values()),SRC_PO+" — net of retail already billed in Wafeq"); r+=1
saas_row=month_row(w1,r,"SaaS subscriptions (Salla App Store, gross)",lambda m: saas_rev.get(m,0.0),SRC_SALLA); r+=1
for s in SAAS_SUBS:
    if s[7]>0:
        month_row(w1,r,f"    {s[0]} — annual, paid {mlbl(s[4][:7])}",lambda m,ss=s: ss[7] if m==ss[4][:7] else 0.0,SRC_SALLA,lvl=1); r+=1
share_row=month_row(w1,r,"less: Salla share (15%)",lambda m: -saas_rev.get(m,0.0)*SALLA_SHARE,"15% per Tarik — not pullable from Salla"); r+=1
tot_all=month_row(w1,r,"TOTAL REVENUE",lambda m: sum(billed[(k,m)] for k in allk)+sum(d.get(m,0.0) for d in acc.values())+saas_rev.get(m,0.0)*(1-SALLA_SHARE),"Billed + to-invoice + SaaS net of Salla share",BOLD,GREY); r+=2
put(w1,r,2,"BY CLIENT (billed)",BOLD); r+=1
for k,cons,st_,chn,brs in CL:
    lbl=k+(f"  (churned {chn})" if chn else "")
    month_row(w1,r,lbl,lambda m,kk=k: billed[(kk,m)],SRC_WAFEQ); r+=1
tot_clients=month_row(w1,r,"TOTAL",lambda m: sum(billed[(k,m)] for k in allk),SRC_WAFEQ,BOLD,GREY); r+=1
assert abs(tot_billed-tot_clients)<0.02 and abs(sum(sums.values())-tot_billed)<0.02
assert abs(tot_all-tot_billed-acc_total-saas_total*(1-SALLA_SHARE))<0.02

# ---------- 2 Brand-Channel Detail (merged, grouped) ----------
w2=wb.create_sheet("2 Brand-Channel Detail")
hdr(w2,["Client","Brand","Channel","Metric","Model","Rate"]+ML+["Total","Source / Notes"],
    widths=[16,17,14,17,7,8]+[10]*len(M)+[11,50])
NC=7+len(M)+1
r=2
NOTES={("Wadi Halfa","fee"):"Temporary support discount (supply issues); card fee 2,000",
       ("Reetal","fee"):"Card 5% + 1,000; billed 4% + 2,000 under setup-period arrangement",
       ("Rimath","fee"):"Card 1,500 = 500 × 3 brands",
       ("Dar Sonbol","fee"):"Card: FAM 2% + 5,000/mo — 50% at 1st live managed channel, 100% from 2nd (FAM only). Salla SaaS = separate added deal.",
       ("Nokush","fee"):"Card: 7%, no monthly fee"}
CARD={"Two United":"Card: MP 2% / Retail 1.5% + 5,000/mo","Ancy & Shaya":"Card: 4% + 1,500/mo",
      "Safwat Aljouf":"Card: 6%","Gamesir":"Card: 1.5%","Inivita":"Card: 7% + 1,500/mo",
      "Wadi Halfa":"Card: MP 4.5% / Retail 4% + 2,000/mo","Alfaris Group":"Card: 6% + 2,500/mo (Salla = SaaS)"}
def mrow(r,cells,vals,src,font=BLACK,fill=None,lvl=0,fmt=SAR):
    for i,v in enumerate(cells,1): put(w2,r,i,v,font,fill=fill)
    tot=0.0
    for j,m in enumerate(M):
        v=vals(m); tot+=v
        put(w2,r,7+j,round(v,2) if abs(v)>0.005 else None,font,fmt,fill)
    put(w2,r,7+len(M),round(tot,2),font,fmt,fill)
    put(w2,r,NC,src,BLACK)
    olvl(w2,r,lvl)
    return tot
grand=defaultdict(float)
for k,cons,st_,chn,brs in CL:
    haseco=k in tree or any(kk==k for (kk,b,ch) in alloc) or any(abs(fee[(k,m)])+abs(setup[(k,m)])+abs(other[(k,m)])+abs(disc[(k,m)])>0.005 for m in M)
    if not haseco: continue
    lbl=k+(f"  (churned {chn})" if chn else "")
    mrow(r,[lbl,None,None,"Billed revenue",None,None],lambda m,kk=k: billed[(kk,m)],SRC_WAFEQ+((" · "+CARD[k]) if k in CARD else ""),BOLD,BILLED_F)
    for j,m in enumerate(M): grand[m]+=billed[(k,m)]
    r+=1
    kb=tree.get(k,{})
    brands=sorted(set(list(kb.keys())+[b for (kk,b,ch) in alloc if kk==k and not b.startswith('—')]+[b for (kk,b,ch) in acc if kk==k]))
    single_brand=len(brands)==1
    pend={(bb,ch):d for (kk,bb,ch),d in alloc.items() if kk==k and not bb.startswith('—')}
    for b in brands:
        chans=sorted(set(list(kb.get(b,{}).keys())+[ch for (bb,ch) in pend if bb==b]+[ch for (kk,bb,ch) in acc if kk==k and bb==b]))
        blvl=1
        if not single_brand:
            mrow(r,[k,disp(b),None,"Sales",None,None],lambda m,bb=b,kk=k: sum(tree[kk][bb][cc].get(m,0.0) for cc in tree[kk].get(bb,{})),"Sum of channel sales below",BLACK,SALES_F,lvl=1)
            r+=1; blvl=2
        for c in chans:
            wch=P2W.get(c,c)
            has_sales=c in kb.get(b,{})
            if has_sales:
                mdl=model(b,c)
                mrow(r,[k,disp(b),c,"Sales",mdl,"0%" if mdl=="SaaS" else None],lambda m,bb=b,cc=c,kk=k: tree[kk][bb][cc].get(m,0.0),tree_src.get((k,b,c),SRC_PLAT),BLACK,SALES_F,lvl=blvl)
                r+=1
            key=(b,c) if (b,c) in pend else ((b,wch) if (b,wch) in pend else None)
            am=pend.pop(key) if key else None
            if am:
                ach=key[1]
                rs=rate_str(k,ach) or ({"Alfaris Group":"6%"}.get(k,"") if ach=='Amazon Retail' else "")
                src=SRC_WAFEQ+(" · brand split pro-rata platform GMV" if (k,ach) in alloc_note else "")
                mrow(r,[k,disp(b)+(" ≈" if (k,ach) in alloc_note else ""),ach,"Commission (billed)",model(b,ach),rs],lambda m,d=am: d.get(m,0.0),src,BLACK,BILLED_F,lvl=blvl)
                r+=1
            if (k,b,c) in acc:
                mrow(r,[k,disp(b),c,"Commission (to invoice)","FAM",f"{acc_rate[(k,b,c)]:g}%"],lambda m,d=acc[(k,b,c)]: d.get(m,0.0),SRC_PO+" — net of billed; on received value",BLACK,TOINV_F,lvl=blvl)
                r+=1
    for (kk,b,ch) in sorted(alloc.keys()):
        if kk==k and b.startswith('—'):
            rs=rate_str(k,ch) or ({"Alfaris Group":"6%"}.get(k,"") if ch=='Amazon Retail' else "")
            mrow(r,[k,b,ch,"Commission (billed)","FAM",rs],lambda m,d=alloc[(kk,b,ch)]: d.get(m,0.0),SRC_WAFEQ,BLACK,BILLED_F,lvl=1)
            r+=1
    if k=="Dar Sonbol":
        for chn2,note2 in [("Amazon","in setup, channel being built"),("Noon","in setup, channel being built"),
                           ("Trendyol","in setup, channel being built"),("aivi","integration WIP")]:
            mrow(r,[k,disp(SONBOL),chn2,"Mapping","FAM","2%"],lambda m: 0.0,SRC_CARD+" — "+note2,BLACK,lvl=1)
            r+=1
    for label,src_,notekey in [("Monthly fee",fee,"fee"),("Setup / one-time",setup,None),("Other",other,None),("Discounts",disc,None)]:
        vals={m:src_[(k,m)] for m in M if abs(src_[(k,m)])>0.005}
        if not vals:
            if label=="Monthly fee" and k in ("Dar Sonbol","Nokush"):
                note=SRC_CARD+" · "+NOTES[(k,"fee")]
                mrow(r,[k,"—",None,label,None,"0" if k=="Nokush" else "—"],lambda m: 0.0,note,BLACK,lvl=1)
                r+=1
            continue
        note=SRC_WAFEQ
        if notekey and (k,notekey) in NOTES: note+=" · "+NOTES[(k,notekey)]
        if label=="Discounts": note+=" · setup-period 50% AM fee / agreed reduction / pricing error"
        mrow(r,[k,"—",None,label,None,None],lambda m,d=dict(vals): d.get(m,0.0),note,BLACK,BILLED_F,lvl=1)
        r+=1
    r+=1
mrow(r,["ALL CLIENTS",None,None,"Billed revenue",None,None],lambda m: grand[m],SRC_WAFEQ,BOLD,GREY); r+=1
mrow(r,["ALL CLIENTS",None,None,"Commission (to invoice)",None,None],lambda m: sum(d.get(m,0.0) for d in acc.values()),SRC_PO,BLACK,TOINV_F); r+=1
mrow(r,["ALL CLIENTS",None,None,"TOTAL incl. to-invoice",None,None],lambda m: grand[m]+sum(d.get(m,0.0) for d in acc.values()),"Billed + closed-PO commission",BOLD,GREY); r+=1

# ---------- 3 SaaS Subscriptions ----------
w3=wb.create_sheet("3 SaaS Subscriptions")
hdr(w3,["Store","Client","Merchant no.","Plan","Start","End","List price","Net (ex-VAT)","Paid (inc-VAT)","Coupon","Status","Source"],
    widths=[22,14,12,9,11,11,10,11,11,13,17,24])
r=2
for s in SAAS_SUBS:
    for i,v in enumerate(s,1):
        c=put(w3,r,i,v)
        if i in (7,8,9) and isinstance(v,(int,float)): c.number_format=SAR
    put(w3,r,12,SRC_SALLA)
    r+=1
put(w3,r,1,"TOTAL PAID (ex-VAT)",BOLD)
put(w3,r,8,round(saas_total,2),BOLD,SAR,GREY)
put(w3,r,12,SRC_SALLA)

OUT=os.path.join(HERE,"output","Nasam_Revenue_View.xlsx")
wb.save(OUT)
print('SAVED', os.path.getsize(OUT), 'bytes')
print('billed:', round(sum(billed.values()),2), '| to-invoice:', round(acc_total,2), '| SaaS subs:', round(saas_total,2),
      '| TOTAL:', round(sum(billed.values())+acc_total+saas_total,2))
print('sales: managed', round(man_tot), '+ retail', round(ret_tot,2), '+ SaaS', round(saas_tot), '=', round(all_tot))
