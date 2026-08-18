#!/usr/bin/env python3
"""Build the final XLSX workbook from eligible.csv / not_eligible.csv."""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = "Tabby_Tamara_KSA_Sellers_for_Nasam.xlsx"
ARIAL = Font(name="Arial", size=10)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="1F4E5F")
TITLE_FONT = Font(name="Arial", size=14, bold=True)
SUB_FONT = Font(name="Arial", size=10, italic=True, color="555555")

WIDTHS = {
    "Store Name (EN)": 32, "Store Name (AR)": 30, "Reason Not Eligible": 45,
    "Website": 40, "Emails": 34, "Phone Numbers": 26, "WhatsApp": 16,
    "Instagram": 34, "Twitter/X": 14, "TikTok": 14, "Snapchat": 14,
    "Categories": 28, "Online Presence": 26, "On Tabby": 9, "On Tamara": 10,
    "Tabby Directory Link": 50, "BNPL Services": 22, "Website Status": 14,
    "Description": 50,
}

def load(path):
    with open(path, encoding="utf-8-sig") as f:
        rdr = csv.reader(f)
        rows = list(rdr)
    return rows[0], rows[1:]

def write_sheet(wb, title, header, rows):
    ws = wb.create_sheet(title)
    ws.append(header)
    for c in ws[1]:
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([v if v != "" else None for v in r])
    for j, h in enumerate(header, 1):
        ws.column_dimensions[get_column_letter(j)].width = WIDTHS.get(h, 18)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            if c.value is not None:
                c.font = ARIAL
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows)+1}"
    return ws

def main():
    h_ok, ok = load("eligible.csv")
    h_no, no = load("not_eligible.csv")

    wb = Workbook()
    smry = wb.active
    smry.title = "Summary"
    el = write_sheet(wb, "Eligible for Nasam", h_ok, ok)
    ne = write_sheet(wb, "Not Eligible", h_no, no)

    col = {name: get_column_letter(i + 1) for i, name in enumerate(h_ok)}
    ncol = {name: get_column_letter(i + 1) for i, name in enumerate(h_no)}

    s = smry
    s["A1"] = "Tabby & Tamara (KSA) — Seller Directory Scrape for Nasam"
    s["A1"].font = TITLE_FONT
    s["A2"] = ("Sources: https://tabby.sa/en-SA/shop (Tabby merchant directory, Saudi Arabia) and "
               "https://tamara.co/ar-sa/stores (Tamara stores directory, Saudi Arabia). "
               "Scraped on 2026-08-18. Deduplicated across both platforms by website domain and normalized store name.")
    s["A2"].font = SUB_FONT

    stats = [
        ("Metric", "Count", None),
        ("Unique sellers — Eligible for Nasam", f"=COUNTA('Eligible for Nasam'!{col['Online Presence']}:{col['Online Presence']})-1", None),
        ("Unique sellers — Not eligible (separate sheet)", f"=COUNTA('Not Eligible'!{ncol['Online Presence']}:{ncol['Online Presence']})-1", None),
        ("Total unique sellers", "=B6+B7", None),
        ("Listed on Tabby", f"=COUNTIF('Eligible for Nasam'!{col['On Tabby']}:{col['On Tabby']},\"Yes\")+COUNTIF('Not Eligible'!{ncol['On Tabby']}:{ncol['On Tabby']},\"Yes\")", None),
        ("Listed on Tamara", f"=COUNTIF('Eligible for Nasam'!{col['On Tamara']}:{col['On Tamara']},\"Yes\")+COUNTIF('Not Eligible'!{ncol['On Tamara']}:{ncol['On Tamara']},\"Yes\")", None),
        ("Listed on both platforms", f"=COUNTIFS('Eligible for Nasam'!{col['On Tabby']}:{col['On Tabby']},\"Yes\",'Eligible for Nasam'!{col['On Tamara']}:{col['On Tamara']},\"Yes\")+COUNTIFS('Not Eligible'!{ncol['On Tabby']}:{ncol['On Tabby']},\"Yes\",'Not Eligible'!{ncol['On Tamara']}:{ncol['On Tamara']},\"Yes\")", None),
        ("Eligible sellers with a website", f"=COUNTA('Eligible for Nasam'!{col['Website']}:{col['Website']})-1", None),
        ("Eligible sellers with email found", f"=COUNTA('Eligible for Nasam'!{col['Emails']}:{col['Emails']})-1", None),
        ("Eligible sellers with phone/WhatsApp found", f"=COUNTA('Eligible for Nasam'!{col['Phone Numbers']}:{col['Phone Numbers']})+COUNTA('Eligible for Nasam'!{col['WhatsApp']}:{col['WhatsApp']})-2", "Counts columns independently; a seller with both is counted in each column."),
    ]
    s["A4"] = "Key figures"
    s["A4"].font = Font(name="Arial", size=12, bold=True)
    r0 = 5
    for i, (label, val, note) in enumerate(stats):
        r = r0 + i
        if i == 0:
            s.cell(row=r, column=1, value=label).font = HDR_FONT
            s.cell(row=r, column=1).fill = HDR_FILL
            s.cell(row=r, column=2, value=val).font = HDR_FONT
            s.cell(row=r, column=2).fill = HDR_FILL
        else:
            s.cell(row=r, column=1, value=label).font = ARIAL
            s.cell(row=r, column=2, value=val).font = ARIAL
            if note:
                s.cell(row=r, column=3, value=note).font = SUB_FONT
    # fix: header consumed one row; stats formulas above reference B5/B6 for first two data rows
    crit_r = r0 + len(stats) + 1
    s.cell(row=crit_r, column=1, value="Eligibility criteria used").font = Font(name="Arial", size=12, bold=True)
    criteria = [
        "ELIGIBLE — product retail businesses that Nasam can onboard and manage across Saudi e-commerce marketplaces:",
        "  • Sells physical consumer products (fashion, beauty, electronics, home, automotive parts, toys, groceries, jewellery, pets, etc.)",
        "  • Independent brand/retailer (not a marketplace aggregator)",
        "  • Includes in-store-only product retailers — flagged 'In-store only' in Online Presence; these are prime prospects to move online",
        "NOT ELIGIBLE — kept in the separate 'Not Eligible' sheet, with the reason per row:",
        "  • Marketplace platforms / global giants (Amazon, Noon, Trendyol, AliExpress, Temu, SHEIN, IKEA, Carrefour, airlines/OTAs, ...)",
        "  • Pure service businesses: clinics, salons & spas, education/training, travel & tourism, restaurants/meal plans,",
        "    insurance/financial services, entertainment & leisure, personal & professional services",
        "  • A seller with BOTH product and service categories is kept in the Eligible sheet (they sell products too).",
        "Contact info (emails, phones, WhatsApp, social handles) was extracted from each seller's own website (homepage + contact page).",
        "'Website Status' shows the crawl result: ok = site reachable; http_4xx/5xx or error = site unreachable at scrape time.",
    ]
    for i, line in enumerate(criteria):
        s.cell(row=crit_r + 1 + i, column=1, value=line).font = ARIAL
    s.column_dimensions["A"].width = 118
    s.column_dimensions["B"].width = 14
    s.column_dimensions["C"].width = 60

    wb.save(OUT)
    print(f"saved {OUT}: eligible={len(ok)}, not_eligible={len(no)}")

if __name__ == "__main__":
    main()
